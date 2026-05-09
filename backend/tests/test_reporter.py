"""
test_reporter.py — Phase 4 Excel Report + Email Alert 測試

=== 測試範圍 ===
  generate_excel_report() — 多工作表 Excel 產生
  send_anomaly_alert()    — Dry run 模式的警報輸出

=== 測試策略 ===
  Excel 測試：把 Workbook 存到 tmp_path（pytest 內建暫存目錄），
              再用 openpyxl.load_workbook() 讀回來驗證結構與內容。
              目的是測「有沒有寫出正確的工作表和資料」，而非格式美不美。

  Email 測試：使用 dry_run=True，讓函式把警報「印到 stdout」而非真的送出。
              再用 capsys 抓 stdout 驗證內容。
              這樣不需要 SMTP 設定，CI 就能跑。

=== Fixture 設計 ===
  make_report()  — 建立最小可用的 ReconciliationReport（帶帳戶、發票、異常）
                   balanced=True/False 控制試算表是否平衡
                   anomaly_count 控制加入幾筆 CRITICAL 異常
"""
import pytest
from pathlib import Path

from openpyxl import load_workbook

from src.reporter import generate_excel_report, send_anomaly_alert
from src.models import (
    AnomalyFlag, ClassifiedInvoice, InvoiceData,
    JournalEntry, LineItem, ReconciliationReport,
)


# ── Test Fixtures ─────────────────────────────────────────────────────────────

def _make_invoice(number="INV-001", vendor="Acme Tech", total=1000.0) -> InvoiceData:
    return InvoiceData(
        invoice_number=number, vendor_name=vendor,
        invoice_date="2024-01-15", due_date=None,
        line_items=[LineItem("Item", 1, total, total)],
        subtotal=total, tax_amount=0.0, total_amount=total,
        currency="USD", source_file="test.pdf",
        extraction_method="pdf_text", raw_text=None,
        extraction_confidence="high",
    )


def _make_ci(number="INV-001", vendor="Acme Tech", total=1000.0,
             confidence="high", category="equipment") -> ClassifiedInvoice:
    inv = _make_invoice(number, vendor, total)
    entries = [
        JournalEntry("Equipment", "Expense", total, 0.0,   f"Debit {number}"),
        JournalEntry("Cash",      "Asset",   0.0,   total, f"Credit {number}"),
    ]
    return ClassifiedInvoice(
        invoice=inv, expense_category=category, expense_subcategory="general",
        classification_confidence=confidence, classification_source="ai",
        journal_entries=entries, ai_reasoning="test",
    )


def _make_flag(severity="critical", inv_number="INV-001", vendor="Acme Tech") -> AnomalyFlag:
    return AnomalyFlag(
        severity=severity, anomaly_type="duplicate_invoice",
        invoice_number=inv_number, vendor_name=vendor,
        description=f"Duplicate Invoice: {vendor} / {inv_number} appears more than once.",
        amount=1000.0,
    )


def make_report(
    *,
    invoice_count: int = 2,
    anomaly_count: int = 0,
    balanced: bool = True,
) -> ReconciliationReport:
    """
    ReconciliationReport 工廠函式。

    invoice_count  — 加入幾張已分類發票
    anomaly_count  — 加入幾筆 CRITICAL 異常（用來測試 Anomaly Report 工作表）
    balanced       — True 時 total_debits == total_credits，False 時差 $1
    """
    invoices = [_make_ci(f"INV-{i:03d}", total=float(i * 500)) for i in range(1, invoice_count + 1)]

    trial_balance = {
        "Equipment": {"debit": 1500.0, "credit": 0.0,    "type": "Expense"},
        "Cash":      {"debit": 0.0,    "credit": 1500.0, "type": "Asset"},
    }
    total_debits  = 1500.0
    total_credits = 1500.0 if balanced else 1499.0

    anomalies = [_make_flag("critical", f"INV-{i:03d}") for i in range(anomaly_count)]

    return ReconciliationReport(
        trial_balance=trial_balance,
        total_debits=total_debits,
        total_credits=total_credits,
        is_balanced=balanced,
        anomalies=anomalies,
        processed_invoices=invoices,
        processing_date="2024-01-15",
        summary_stats={
            "invoice_count": invoice_count,
            "total_amount": sum(i * 500.0 for i in range(1, invoice_count + 1)),
            "by_category": {"equipment": {"count": invoice_count, "total": 1500.0}},
            "by_source":   {"ai": invoice_count},
            "anomaly_counts": {"critical": anomaly_count},
        },
    )


# ── generate_excel_report() ───────────────────────────────────────────────────

class TestGenerateExcelReport:
    def test_creates_file_at_given_path(self, tmp_path):
        # 最基本：函式必須在指定路徑建立檔案
        report = make_report()
        out = tmp_path / "test_report.xlsx"
        result = generate_excel_report(report, out)
        assert result == out
        assert out.exists()

    def test_creates_file_at_default_path_when_none(self, tmp_path, monkeypatch):
        # output_path=None 時，應使用 config.OUTPUT_DIR / "reconciliation_{date}.xlsx"
        import src.reporter as reporter_mod
        import config
        monkeypatch.setattr(config, "OUTPUT_DIR", tmp_path)
        monkeypatch.setattr(reporter_mod, "OUTPUT_DIR", tmp_path)
        report = make_report()
        path = generate_excel_report(report)
        assert path.exists()
        assert "reconciliation_2024-01-15" in path.name

    def test_three_sheets_when_no_anomalies(self, tmp_path):
        # 沒有異常 → 不產生 Anomaly Report 工作表，只有 3 張
        report = make_report(anomaly_count=0)
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        assert set(wb.sheetnames) == {"Trial Balance", "Income Statement", "Invoice Detail"}

    def test_four_sheets_when_anomalies_present(self, tmp_path):
        # 有異常 → 4 張工作表，含 Anomaly Report
        report = make_report(anomaly_count=2)
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        assert "Anomaly Report" in wb.sheetnames

    # ── Trial Balance 工作表 ─────────────────────────────────────────────────

    def test_trial_balance_headers(self, tmp_path):
        # 第 1 行應該是欄位標題
        report = make_report()
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Trial Balance"]
        headers = [ws.cell(1, col).value for col in range(1, 6)]
        assert headers == [
            "Account Name", "Account Type", "Total Debits", "Total Credits", "Net Balance",
        ]

    def test_trial_balance_contains_account_data(self, tmp_path):
        # trial_balance 裡的帳戶名稱必須出現在工作表的資料行
        report = make_report()
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Trial Balance"]
        account_names = {ws.cell(row, 1).value for row in range(2, ws.max_row + 1)}
        assert "Equipment" in account_names
        assert "Cash" in account_names

    def test_trial_balance_totals_row(self, tmp_path):
        # TOTAL 行應出現在最後一筆帳戶之後
        report = make_report()
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Trial Balance"]
        all_values = [ws.cell(row, 1).value for row in range(1, ws.max_row + 1)]
        assert "TOTAL" in all_values

    def test_trial_balance_balanced_status(self, tmp_path):
        # 平衡時應出現「BALANCED」字樣
        report = make_report(balanced=True)
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Trial Balance"]
        all_text = " ".join(str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1))
        assert "BALANCED" in all_text

    # ── Income Statement 工作表 ───────────────────────────────────────────────

    def test_income_statement_has_title(self, tmp_path):
        # 第 1 行是含日期的標題
        report = make_report()
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Income Statement"]
        title = ws.cell(1, 1).value
        assert title is not None
        assert "2024-01-15" in title

    def test_income_statement_has_expense_section(self, tmp_path):
        # Expense 帳戶應出現在 Income Statement 的 EXPENSES 段落
        report = make_report()
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Income Statement"]
        all_text = " ".join(str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1))
        assert "EXPENSES" in all_text
        assert "Equipment" in all_text    # trial_balance 裡有 Equipment / Expense

    # ── Invoice Detail 工作表 ─────────────────────────────────────────────────

    def test_invoice_detail_headers(self, tmp_path):
        report = make_report()
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Invoice Detail"]
        headers = [ws.cell(1, col).value for col in range(1, 9)]
        assert "Invoice Number" in headers
        assert "Vendor" in headers
        assert "Amount" in headers

    def test_invoice_detail_row_count(self, tmp_path):
        # 發票數量 + 1 個 header = max_row（至少）
        report = make_report(invoice_count=3)
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Invoice Detail"]
        data_rows = [
            ws.cell(row, 1).value
            for row in range(2, ws.max_row + 1)
            if ws.cell(row, 1).value is not None
        ]
        assert len(data_rows) == 3

    def test_invoice_detail_confidence_column(self, tmp_path):
        # Confidence 欄（第 5 欄）應出現在資料行
        report = make_report(invoice_count=1)
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Invoice Detail"]
        assert ws.cell(2, 5).value in ("high", "medium", "low")

    # ── Anomaly Report 工作表 ─────────────────────────────────────────────────

    def test_anomaly_report_headers(self, tmp_path):
        report = make_report(anomaly_count=1)
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Anomaly Report"]
        headers = [ws.cell(1, col).value for col in range(1, 7)]
        assert "Severity" in headers
        assert "Description" in headers

    def test_anomaly_report_row_count(self, tmp_path):
        # 1 個 header + N 筆異常 = max_row
        report = make_report(anomaly_count=3)
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Anomaly Report"]
        data_rows = [
            ws.cell(row, 1).value
            for row in range(2, ws.max_row + 1)
            if ws.cell(row, 1).value is not None
        ]
        assert len(data_rows) == 3

    def test_anomaly_severity_uppercased(self, tmp_path):
        # Severity 欄應顯示大寫（CRITICAL / WARNING / INFO）
        report = make_report(anomaly_count=1)
        wb = load_workbook(generate_excel_report(report, tmp_path / "r.xlsx"))
        ws = wb["Anomaly Report"]
        assert ws.cell(2, 1).value == "CRITICAL"


# ── send_anomaly_alert() ──────────────────────────────────────────────────────

class TestSendAnomalyAlert:
    def test_returns_false_when_no_anomalies(self, tmp_path):
        # 沒有 CRITICAL 或 WARNING → 不需要警報，回傳 False
        report = make_report(anomaly_count=0)
        path = tmp_path / "r.xlsx"
        path.touch()
        assert send_anomaly_alert(report, path, dry_run=True) is False

    def test_returns_true_for_critical_anomaly(self, tmp_path, capsys):
        # 有 CRITICAL 異常 → dry_run 印出警報，回傳 True
        report = make_report(anomaly_count=1)
        path = tmp_path / "r.xlsx"
        path.touch()
        result = send_anomaly_alert(report, path, dry_run=True)
        assert result is True

    def test_dry_run_prints_to_stdout(self, tmp_path, capsys):
        # dry_run 模式下，警報內容應印到 stdout（不送出真實 email）
        report = make_report(anomaly_count=1)
        path = tmp_path / "r.xlsx"
        path.touch()
        send_anomaly_alert(report, path, dry_run=True)
        captured = capsys.readouterr()
        assert "DRY RUN" in captured.out

    def test_dry_run_output_contains_report_path(self, tmp_path, capsys):
        # 警報內容應包含 Excel 報表的路徑（讓收到警報的人知道去哪看）
        report = make_report(anomaly_count=1)
        path = tmp_path / "report.xlsx"
        path.touch()
        send_anomaly_alert(report, path, dry_run=True)
        captured = capsys.readouterr()
        assert "report.xlsx" in captured.out

    def test_warning_only_also_triggers_alert(self, tmp_path, capsys):
        # WARNING 等級的異常（如大額發票）也應該觸發警報，雖然沒有 CRITICAL
        warning_flag = AnomalyFlag(
            severity="warning", anomaly_type="large_amount",
            invoice_number="INV-001", vendor_name="Acme Tech",
            description="Large amount $15,000 exceeds threshold.", amount=15000.0,
        )
        report = make_report(anomaly_count=0)
        report.anomalies.append(warning_flag)
        path = tmp_path / "r.xlsx"
        path.touch()
        result = send_anomaly_alert(report, path, dry_run=True)
        assert result is True

    def test_info_only_does_not_trigger_alert(self, tmp_path):
        # INFO 等級（如舊發票）不值得驚動財務人員，不發警報
        info_flag = AnomalyFlag(
            severity="info", anomaly_type="old_invoice",
            invoice_number="INV-001", vendor_name="Acme Tech",
            description="Invoice is 95 days old.", amount=500.0,
        )
        report = make_report(anomaly_count=0)
        report.anomalies.append(info_flag)
        path = tmp_path / "r.xlsx"
        path.touch()
        assert send_anomaly_alert(report, path, dry_run=True) is False
