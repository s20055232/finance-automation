"""
reporter.py — Phase 4: Excel Report Generation + Email Alert

Four-sheet workbook structure:
  1. Trial Balance    — Account rollup, Debit/Credit totals, Net Balance
  2. Income Statement — Revenue vs Expenses, Net Income/Loss, % of Revenue
  3. Invoice Detail   — One row per invoice, confidence color-coded
  4. Anomaly Report   — Only rendered when anomalies exist, severity color-coded
"""

import logging
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    ALERT_EMAIL_TO, EMAIL_DRY_RUN, OUTPUT_DIR,
    SMTP_HOST, SMTP_PASSWORD, SMTP_PORT, SMTP_USERNAME,
)
from src.models import AnomalyFlag, ReconciliationReport

logger = logging.getLogger(__name__)


# ── Color / Style Constants ────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy blue
_ALT_ROW_FILL = PatternFill("solid", fgColor="D9E1F2")   # light periwinkle
_TOTAL_FILL   = PatternFill("solid", fgColor="BDD7EE")   # medium blue
_SECTION_FILL = PatternFill("solid", fgColor="2E75B6")   # section header blue

_CRITICAL_FILL = PatternFill("solid", fgColor="FFCCCC")  # light red
_WARNING_FILL  = PatternFill("solid", fgColor="FFE0B2")  # light orange
_INFO_FILL     = PatternFill("solid", fgColor="E3F2FD")  # light blue

_HIGH_FILL = PatternFill("solid", fgColor="C8E6C9")      # green  — high confidence
_MED_FILL  = PatternFill("solid", fgColor="FFF9C4")      # yellow — medium confidence
_LOW_FILL  = PatternFill("solid", fgColor="FFCCCC")      # red    — low confidence

_HEADER_FONT  = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
_WHITE_FONT   = Font(name="Calibri", color="FFFFFF", bold=True)
_BOLD_FONT    = Font(name="Calibri", bold=True)
_RED_FONT     = Font(name="Calibri", color="C00000", bold=True)
_GREEN_FONT   = Font(name="Calibri", color="006100", bold=True)

_CURRENCY_FMT = '$#,##0.00'
_PCT_FMT      = '0.0%'
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT  = Alignment(horizontal="right",  vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")

_SEVERITY_FILL = {"critical": _CRITICAL_FILL, "warning": _WARNING_FILL, "info": _INFO_FILL}
_CONF_FILL     = {"high": _HIGH_FILL, "medium": _MED_FILL}


# ── Public API ────────────────────────────────────────────────────────────────

def generate_excel_report(
    report: ReconciliationReport,
    output_path: Optional[Path] = None,
) -> Path:
    """
    Generate a multi-sheet Excel workbook from a ReconciliationReport.
    Returns the path to the saved file.
    """
    if output_path is None:
        output_path = OUTPUT_DIR / f"reconciliation_{report.processing_date}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # discard the default empty sheet

    _write_trial_balance(wb, report)
    _write_income_statement(wb, report)
    _write_invoice_detail(wb, report)
    if report.anomalies:
        _write_anomaly_report(wb, report)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Excel report saved to %s (%d sheets)", output_path, len(wb.sheetnames))
    return output_path


def send_anomaly_alert(
    report: ReconciliationReport,
    report_path: Path,
    *,
    dry_run: Optional[bool] = None,
) -> bool:
    """
    Send an email alert when CRITICAL or WARNING anomalies are present.
    Returns True if the alert was sent (or printed in dry_run mode).
    Returns False when there is nothing worth alerting.
    """
    if dry_run is None:
        dry_run = EMAIL_DRY_RUN

    critical = [f for f in report.anomalies if f.severity == "critical"]
    warnings  = [f for f in report.anomalies if f.severity == "warning"]
    if not critical and not warnings:
        return False

    severity_tag = "CRITICAL" if critical else "WARNING"
    subject = (
        f"[Finance Bot] {severity_tag}: "
        f"{len(critical)} critical, {len(warnings)} warning(s) "
        f"— {report.processing_date}"
    )
    body = _build_alert_body(report, critical, warnings, report_path)

    if dry_run:
        print("\n" + "─" * 70)
        print("DRY RUN — email alert (not sent):")
        print(f"  To:      {ALERT_EMAIL_TO or '(ALERT_EMAIL_TO not set)'}")
        print(f"  Subject: {subject}")
        print()
        print(body)
        print("─" * 70 + "\n")
        return True

    if not ALERT_EMAIL_TO:
        logger.warning("ALERT_EMAIL_TO not configured — skipping email alert")
        return False

    return _send_smtp(subject, body, report_path)


# ── Sheet 1: Trial Balance ─────────────────────────────────────────────────────

def _write_trial_balance(wb: Workbook, report: ReconciliationReport) -> None:
    ws = wb.create_sheet("Trial Balance")

    headers = ["Account Name", "Account Type", "Total Debits", "Total Credits", "Net Balance"]
    _write_header_row(ws, 1, headers)

    for row_idx, account in enumerate(sorted(report.trial_balance), start=2):
        data   = report.trial_balance[account]
        debit  = data["debit"]
        credit = data["credit"]
        net    = round(debit - credit, 2)
        fill   = _ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()

        ws.cell(row_idx, 1, account).fill       = fill
        ws.cell(row_idx, 2, data["type"]).fill  = fill

        for col, val in ((3, debit), (4, credit)):
            c = ws.cell(row_idx, col, val)
            c.fill, c.number_format, c.alignment = fill, _CURRENCY_FMT, _RIGHT

        net_cell = ws.cell(row_idx, 5, net)
        net_cell.fill, net_cell.number_format, net_cell.alignment = fill, _CURRENCY_FMT, _RIGHT
        if net < 0:
            net_cell.font = _RED_FONT

    # Totals row
    total_row = len(report.trial_balance) + 2
    ws.cell(total_row, 1, "TOTAL").font = _BOLD_FONT
    for col, val in ((3, report.total_debits), (4, report.total_credits)):
        c = ws.cell(total_row, col, val)
        c.fill, c.number_format, c.alignment, c.font = _TOTAL_FILL, _CURRENCY_FMT, _RIGHT, _BOLD_FONT

    net_total = round(report.total_debits - report.total_credits, 2)
    net_cell = ws.cell(total_row, 5, net_total)
    net_cell.fill, net_cell.number_format, net_cell.alignment = _TOTAL_FILL, _CURRENCY_FMT, _RIGHT
    net_cell.font = _RED_FONT if net_total < 0 else _BOLD_FONT
    for col in (1, 2):
        ws.cell(total_row, col).fill = _TOTAL_FILL

    # Balance status note below totals
    status_row = total_row + 2
    status_text = (
        "✓ Trial Balance is BALANCED"
        if report.is_balanced
        else "✗ Trial Balance is UNBALANCED — review all Journal Entries"
    )
    ws.cell(status_row, 1, status_text).font = _GREEN_FONT if report.is_balanced else _RED_FONT

    _auto_column_widths(ws)
    ws.freeze_panes = "A2"


# ── Sheet 2: Income Statement ──────────────────────────────────────────────────

def _write_income_statement(wb: Workbook, report: ReconciliationReport) -> None:
    ws = wb.create_sheet("Income Statement")

    # Title
    ws.cell(1, 1, f"Income Statement — {report.processing_date}").font = Font(
        name="Calibri", bold=True, size=14, color="1F3864"
    )

    # Column labels
    for col, label in enumerate(["Account", "Amount", "% of Revenue"], start=1):
        cell = ws.cell(3, col, label)
        cell.font, cell.alignment = _BOLD_FONT, (_RIGHT if col > 1 else _LEFT)

    revenue_accts = {k: v for k, v in report.trial_balance.items() if v["type"] == "Revenue"}
    expense_accts = {k: v for k, v in report.trial_balance.items() if v["type"] == "Expense"}
    total_revenue = sum(v["credit"] - v["debit"] for v in revenue_accts.values())

    row = 4
    row = _write_is_section(ws, row, "REVENUE",  revenue_accts, is_revenue=True,  pct_base=total_revenue)
    row += 1
    row = _write_is_section(ws, row, "EXPENSES", expense_accts, is_revenue=False, pct_base=total_revenue)
    row += 1

    total_expenses = sum(v["debit"] - v["credit"] for v in expense_accts.values())
    net = round(total_revenue - total_expenses, 2)
    label = "Net Income" if net >= 0 else "Net Loss"

    ws.cell(row, 1, label).font = _BOLD_FONT
    amt_cell = ws.cell(row, 2, net)
    amt_cell.number_format, amt_cell.alignment = _CURRENCY_FMT, _RIGHT
    amt_cell.font = _GREEN_FONT if net >= 0 else _RED_FONT
    if total_revenue > 0:
        pct_cell = ws.cell(row, 3, net / total_revenue)
        pct_cell.number_format, pct_cell.alignment = _PCT_FMT, _RIGHT
    for col in range(1, 4):
        ws.cell(row, col).fill = _TOTAL_FILL

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16


def _write_is_section(
    ws: Worksheet,
    start_row: int,
    title: str,
    accounts: dict,
    *,
    is_revenue: bool,
    pct_base: float,
) -> int:
    """Write one section (Revenue or Expenses) of the Income Statement. Returns the next row."""
    row = start_row

    # Section header bar
    for col in range(1, 4):
        ws.cell(row, col).fill = _SECTION_FILL
    ws.cell(row, 1, title).font = _WHITE_FONT
    row += 1

    section_total = 0.0
    for name, data in sorted(accounts.items()):
        # Revenue: normal credit balance → credit - debit = positive amount earned
        # Expense: normal debit balance  → debit - credit = positive amount spent
        amount = round(
            (data["credit"] - data["debit"]) if is_revenue else (data["debit"] - data["credit"]),
            2,
        )
        section_total += amount
        ws.cell(row, 1, "  " + name)
        c = ws.cell(row, 2, amount)
        c.number_format, c.alignment = _CURRENCY_FMT, _RIGHT
        if pct_base > 0:
            p = ws.cell(row, 3, amount / pct_base)
            p.number_format, p.alignment = _PCT_FMT, _RIGHT
        row += 1

    # Section subtotal
    subtotal_label = "  Total Revenue" if is_revenue else "  Total Expenses"
    ws.cell(row, 1, subtotal_label).font = _BOLD_FONT
    c = ws.cell(row, 2, round(section_total, 2))
    c.number_format, c.alignment, c.font = _CURRENCY_FMT, _RIGHT, _BOLD_FONT
    if pct_base > 0:
        p = ws.cell(row, 3, section_total / pct_base if pct_base > 0 else None)
        if p.value is not None:
            p.number_format, p.alignment = _PCT_FMT, _RIGHT
    for col in range(1, 4):
        ws.cell(row, col).fill = _ALT_ROW_FILL
    row += 1

    return row


# ── Sheet 3: Invoice Detail ────────────────────────────────────────────────────

def _write_invoice_detail(wb: Workbook, report: ReconciliationReport) -> None:
    ws = wb.create_sheet("Invoice Detail")

    headers = [
        "Invoice Number", "Vendor", "Date", "Category",
        "Confidence", "Source", "Amount", "Anomalies",
    ]
    _write_header_row(ws, 1, headers)

    # Pre-index which invoice numbers have anomalies and at what severity
    anomaly_map: dict[str, list[str]] = {}
    for flag in report.anomalies:
        anomaly_map.setdefault(flag.invoice_number, []).append(flag.severity.upper())

    for row_idx, ci in enumerate(report.processed_invoices, start=2):
        inv  = ci.invoice
        fill = _ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()

        for col, val in (
            (1, inv.invoice_number),
            (2, inv.vendor_name),
            (3, inv.invoice_date),
            (4, ci.expense_category),
            (6, ci.classification_source),
        ):
            ws.cell(row_idx, col, val).fill = fill

        # Confidence cell: green / yellow / red
        conf_cell = ws.cell(row_idx, 5, ci.classification_confidence)
        conf_cell.fill = _CONF_FILL.get(ci.classification_confidence, _LOW_FILL)
        conf_cell.alignment = _CENTER

        amt_cell = ws.cell(row_idx, 7, inv.total_amount)
        amt_cell.number_format, amt_cell.alignment, amt_cell.fill = _CURRENCY_FMT, _RIGHT, fill

        anomaly_labels = anomaly_map.get(inv.invoice_number)
        ws.cell(row_idx, 8, ", ".join(anomaly_labels) if anomaly_labels else "—").fill = fill

    _auto_column_widths(ws)
    ws.freeze_panes = "A2"


# ── Sheet 4: Anomaly Report ────────────────────────────────────────────────────

def _write_anomaly_report(wb: Workbook, report: ReconciliationReport) -> None:
    ws = wb.create_sheet("Anomaly Report")

    headers = ["Severity", "Type", "Invoice Number", "Vendor", "Description", "Amount"]
    _write_header_row(ws, 1, headers)

    for row_idx, flag in enumerate(report.anomalies, start=2):
        fill = _SEVERITY_FILL.get(flag.severity, PatternFill())

        sev_cell = ws.cell(row_idx, 1, flag.severity.upper())
        sev_cell.fill, sev_cell.font, sev_cell.alignment = fill, _BOLD_FONT, _CENTER

        for col, val in (
            (2, flag.anomaly_type),
            (3, flag.invoice_number),
            (4, flag.vendor_name),
        ):
            ws.cell(row_idx, col, val).fill = fill

        desc_cell = ws.cell(row_idx, 5, flag.description)
        desc_cell.fill = fill
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

        if flag.amount is not None:
            amt_cell = ws.cell(row_idx, 6, flag.amount)
            amt_cell.fill, amt_cell.number_format, amt_cell.alignment = fill, _CURRENCY_FMT, _RIGHT
        else:
            ws.cell(row_idx, 6, "—").fill = fill

    ws.column_dimensions["E"].width = 60  # Description needs more room
    _auto_column_widths(ws, skip_cols={5})
    ws.freeze_panes = "A2"


# ── Style Helpers ─────────────────────────────────────────────────────────────

def _write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row, col, label)
        cell.font, cell.fill, cell.alignment = _HEADER_FONT, _HEADER_FILL, _CENTER


def _auto_column_widths(ws: Worksheet, skip_cols: Optional[set] = None) -> None:
    """Set column widths based on maximum content length, capped at 50."""
    skip_cols = skip_cols or set()
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        if col_idx in skip_cols:
            continue
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)


# ── Email ─────────────────────────────────────────────────────────────────────

def _build_alert_body(
    report: ReconciliationReport,
    critical: list[AnomalyFlag],
    warnings: list[AnomalyFlag],
    report_path: Path,
) -> str:
    lines = [
        "Finance Automation — Reconciliation Alert",
        f"Processing Date:    {report.processing_date}",
        f"Invoices Processed: {len(report.processed_invoices)}",
        f"Trial Balance:      {'BALANCED' if report.is_balanced else 'UNBALANCED'}",
        f"Report:             {report_path}",
        "",
    ]
    if critical:
        lines.append(f"CRITICAL ({len(critical)}):")
        for f in critical:
            lines.append(f"  • [{f.invoice_number}] {f.vendor_name}: {f.description}")
        lines.append("")
    if warnings:
        lines.append(f"WARNINGS ({len(warnings)}):")
        for f in warnings:
            lines.append(f"  • [{f.invoice_number}] {f.vendor_name}: {f.description}")
    return "\n".join(lines)


def _send_smtp(subject: str, body: str, attachment_path: Path) -> bool:
    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"]    = SMTP_USERNAME
    msg["To"]      = ALERT_EMAIL_TO
    msg.attach(MIMEText(body, "plain"))

    with open(attachment_path, "rb") as fh:
        part = MIMEApplication(fh.read(), Name=attachment_path.name)
        part["Content-Disposition"] = f'attachment; filename="{attachment_path.name}"'
        msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_USERNAME, ALERT_EMAIL_TO, msg.as_string())
        logger.info("Alert email sent to %s", ALERT_EMAIL_TO)
        return True
    except Exception as exc:
        logger.error("Failed to send alert email: %s", exc)
        return False
